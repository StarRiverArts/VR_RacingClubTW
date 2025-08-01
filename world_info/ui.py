"""Tkinter GUI for fetching and filtering VRChat world data.

The interface provides several tabs:
- Entrance: input authentication (cookie or basic auth), a search keyword
  and creator user ID.
- Data: fetch worlds by keyword and display the raw JSON.
- Filter: filter the keyword results by tag and sort order.
- World List: show the filtered worlds in a simple list.
- User Worlds: fetch and display worlds created by a specific user.

The tool relies on functions in ``scraper/scraper.py``.  Results are saved
under that folder for reuse by other scripts.  Fetching a creator's worlds uses
Playwright to scrape the VRChat website, so the ``playwright`` package must be
installed and ``playwright install`` executed beforehand.
"""
from __future__ import annotations

import json
import datetime as dt
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox

from scraper.scraper import (
    fetch_worlds,
    _load_headers,
    load_history,
    update_history,
    record_row,
    _parse_date,
    EXCEL_FILE,
    HISTORY_TABLE,
)

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional
    load_workbook = None  # type: ignore

BASE = Path(__file__).resolve().parent
RAW_FILE = BASE / "scraper" / "raw_worlds.json"
USER_FILE = BASE / "scraper" / "user_worlds.json"

# Column headers for metrics tables
METRIC_COLS = [
    "世界名稱",
    "世界ID",
    "發布日期",
    "最後更新",
    "瀏覽人次",
    "大小",
    "收藏次數",
    "熱度",
    "人氣",
    "實驗室到發布",
    "瀏覽蒐藏比",
    "距離上次更新",
    "已發布",
    "人次發布比",
]

# Legend for line charts
LEGEND_TEXT = "藍:人次 綠:收藏 紅:熱度 紫:熱門度 橘:實驗室 黑:公開 灰:更新"

class WorldInfoUI(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("World Info")
        self.geometry("800x600")

        self.headers = {}
        self.data: list[dict] = []
        self.user_data: list[dict] = []
        self.filtered: list[dict] = []
        self.history: dict[str, list[dict]] = load_history()

        self._build_tabs()

    # ------------------------------------------------------------------
    # UI construction
    def _build_tabs(self) -> None:
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True)

        self.tab_entry = ttk.Frame(self.nb)
        self.tab_data = ttk.Frame(self.nb)
        self.tab_filter = ttk.Frame(self.nb)
        self.tab_list = ttk.Frame(self.nb)
        self.tab_user = ttk.Frame(self.nb)
        self.tab_history = ttk.Frame(self.nb)

        self.nb.add(self.tab_entry, text="入口")
        self.nb.add(self.tab_data, text="資料")
        self.nb.add(self.tab_filter, text="篩選")
        self.nb.add(self.tab_list, text="世界列表")
        self.nb.add(self.tab_user, text="個人世界")
        self.nb.add(self.tab_history, text="歷史記錄")

        self._build_entry_tab()
        self._build_data_tab()
        self._build_filter_tab()
        self._build_list_tab()
        self._build_user_tab()
        self._build_history_tab()
        self._load_local_tables()

    # ------------------------------------------------------------------
    # Entry tab widgets
    def _build_entry_tab(self) -> None:
        f = self.tab_entry
        row = 0
        ttk.Label(f, text="Cookie").grid(row=row, column=0, sticky="e")
        self.var_cookie = tk.StringVar()
        tk.Entry(f, textvariable=self.var_cookie, width=60).grid(row=row, column=1, padx=4, pady=2)
        row += 1

        ttk.Label(f, text="Username").grid(row=row, column=0, sticky="e")
        self.var_user = tk.StringVar()
        tk.Entry(f, textvariable=self.var_user).grid(row=row, column=1, padx=4, pady=2)
        row += 1

        ttk.Label(f, text="Password").grid(row=row, column=0, sticky="e")
        self.var_pass = tk.StringVar()
        tk.Entry(f, textvariable=self.var_pass, show="*").grid(row=row, column=1, padx=4, pady=2)
        row += 1

        ttk.Label(f, text="Search Keyword").grid(row=row, column=0, sticky="e")
        self.var_keyword = tk.StringVar()
        tk.Entry(f, textvariable=self.var_keyword).grid(row=row, column=1, padx=4, pady=2)
        row += 1

        ttk.Label(f, text="User ID").grid(row=row, column=0, sticky="e")
        self.var_userid = tk.StringVar()
        tk.Entry(f, textvariable=self.var_userid).grid(row=row, column=1, padx=4, pady=2)
        row += 1

        btn_frame = ttk.Frame(f)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=4)
        ttk.Button(btn_frame, text="Search", command=self._on_search).grid(row=0, column=0, padx=2)
        ttk.Button(btn_frame, text="User Worlds", command=self._on_user).grid(row=0, column=1, padx=2)

    # ------------------------------------------------------------------
    # Data tab widgets
    def _build_data_tab(self) -> None:
        self.text_data = tk.Text(self.tab_data, wrap="word")
        self.text_data.pack(fill=tk.BOTH, expand=True)
        ttk.Button(self.tab_data, text="Open Filter", command=lambda: self.nb.select(self.tab_filter)).pack(pady=4)

    # Filter tab widgets
    def _build_filter_tab(self) -> None:
        f = self.tab_filter
        ttk.Label(f, text="Tag").grid(row=0, column=0, sticky="e")
        self.var_tag = tk.StringVar(value="all")
        self.box_tag = ttk.Combobox(f, textvariable=self.var_tag, values=["all"])
        self.box_tag.grid(row=0, column=1, padx=4, pady=2)

        ttk.Label(f, text="Sort").grid(row=1, column=0, sticky="e")
        self.var_sort = tk.StringVar(value="popular")
        ttk.Combobox(f, textvariable=self.var_sort, values=["latest", "popular"]).grid(row=1, column=1, padx=4, pady=2)

        ttk.Button(f, text="Apply", command=self._apply_filter).grid(row=2, column=0, columnspan=2, pady=4)

    # World list tab
    def _build_list_tab(self) -> None:
        columns = ("name", "visits", "id")
        self.tree = ttk.Treeview(self.tab_list, columns=columns, show="headings")
        self.tree.heading("name", text="Name")
        self.tree.heading("visits", text="Visits")
        self.tree.heading("id", text="World ID")
        self.tree.column("name", width=250)
        self.tree.column("visits", width=80, anchor="e")
        self.tree.column("id", width=200)
        vsb = ttk.Scrollbar(self.tab_list, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill=tk.BOTH, expand=True)
        vsb.pack(side="right", fill=tk.Y)

    # User worlds tab
    def _build_user_tab(self) -> None:
        f = self.tab_user
        self.user_nb = ttk.Notebook(f)
        self.user_nb.pack(fill=tk.BOTH, expand=True)

        # dashboard and detail notebooks
        self.tab_dashboard = ttk.Frame(self.user_nb)
        self.tab_detail = ttk.Frame(self.user_nb)
        self.user_nb.add(self.tab_dashboard, text="儀表板")
        self.user_nb.add(self.tab_detail, text="詳細列表")

        self._build_dashboard_tab()

        self.detail_nb = ttk.Notebook(self.tab_detail)
        self.detail_nb.pack(fill=tk.BOTH, expand=True)

        self.tab_user_list = ttk.Frame(self.detail_nb)
        self.detail_nb.add(self.tab_user_list, text="所有世界")

        self.user_tree = ttk.Treeview(self.tab_user_list, show="headings")
        columns = ["爬取日期"] + METRIC_COLS
        self.user_tree["columns"] = list(range(len(columns)))
        for idx, col in enumerate(columns):
            self.user_tree.heading(str(idx), text=col)
            self.user_tree.column(str(idx), width=80, anchor="center")
        vsb = ttk.Scrollbar(self.tab_user_list, orient="vertical", command=self.user_tree.yview)
        self.user_tree.configure(yscrollcommand=vsb.set)
        self.user_tree.pack(side="left", fill=tk.BOTH, expand=True)
        vsb.pack(side="right", fill=tk.Y)
        self.user_tree.bind("<<TreeviewSelect>>", self._on_select_user_world)

        self.user_canvas = tk.Canvas(self.tab_user_list, bg="white", height=200)
        self.user_canvas.pack(fill=tk.BOTH, expand=True)
        ttk.Label(self.tab_user_list, text=LEGEND_TEXT).pack()

    def _build_dashboard_tab(self) -> None:
        """Create the dashboard view with a summary table and charts."""
        f = self.tab_dashboard
        self.dash_tree = ttk.Treeview(f, show="headings")
        self.dash_tree["columns"] = list(range(len(METRIC_COLS)))
        for idx, col in enumerate(METRIC_COLS):
            self.dash_tree.heading(str(idx), text=col)
            self.dash_tree.column(str(idx), width=80, anchor="center")
        self.dash_tree.pack(fill=tk.X)

        self.chart_container = ttk.Frame(f)
        self.chart_container.pack(fill=tk.BOTH, expand=True)
        self.chart_container.bind("<Configure>", self._arrange_dashboard_charts)
        self.chart_frames: list[tuple[tk.Frame, tk.Canvas, dict]] = []

    def _build_history_tab(self) -> None:
        f = self.tab_history
        self.var_hist_world = tk.StringVar()
        self.box_hist_world = ttk.Combobox(f, textvariable=self.var_hist_world, values=list(self.history.keys()))
        self.box_hist_world.pack(fill=tk.X, pady=2)
        self.box_hist_world.bind("<<ComboboxSelected>>", self._draw_history)
        self.canvas = tk.Canvas(f, bg="white")
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self._update_history_options()

    def _load_local_tables(self) -> None:
        """Load existing Excel history and populate the user world list."""
        if load_workbook is None:
            return
        if EXCEL_FILE.exists():
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.user_tree.insert("", tk.END, values=row)
                if len(row) == 15:
                    (
                        fetched,
                        name,
                        wid,
                        pub,
                        upd,
                        visits,
                        size,
                        fav,
                        heat,
                        pop,
                        labs_to_pub,
                        vf,
                        since_upd,
                        released,
                        vpp,
                    ) = row
                else:
                    # backward compatibility with old files without fetch date
                    fetched = ""
                    (
                        name,
                        wid,
                        pub,
                        upd,
                        visits,
                        size,
                        fav,
                        heat,
                        pop,
                        labs_to_pub,
                        vf,
                        since_upd,
                        released,
                        vpp,
                    ) = row
                self.user_data.append(
                    {
                        "爬取日期": fetched,
                        "世界名稱": name,
                        "世界ID": wid,
                        "發布日期": pub,
                        "最後更新": upd,
                        "瀏覽人次": visits,
                        "大小": size,
                        "收藏次數": fav,
                        "熱度": heat,
                        "人氣": pop,
                        "實驗室到發布": labs_to_pub,
                        "瀏覽蒐藏比": vf,
                        "距離上次更新": since_upd,
                        "已發布": released,
                        "人次發布比": vpp,
                    }
                )
            self._create_world_tabs()
            self._update_dashboard()
    # ------------------------------------------------------------------
    # Actions
    def _load_auth_headers(self) -> None:
        cookie = self.var_cookie.get() or None
        user = self.var_user.get() or None
        pw = self.var_pass.get() or None
        self.headers = _load_headers(cookie, user, pw)

    def _on_search(self) -> None:
        self._load_auth_headers()
        keyword = self.var_keyword.get().strip()
        if not keyword:
            messagebox.showerror("Error", "Keyword required")
            return
        try:
            self.data = fetch_worlds(keyword=keyword, limit=50, headers=self.headers)
            with open(RAW_FILE, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            update_history(self.data)
            self.history = load_history()
            self._update_history_options()
            self.text_data.delete("1.0", tk.END)
            self.text_data.insert(tk.END, json.dumps(self.data, ensure_ascii=False, indent=2))
            self._update_tag_options()
            self.nb.select(self.tab_data)
        except RuntimeError as e:  # pragma: no cover - runtime only
            messagebox.showerror("HTTP Error", str(e))
        except Exception as e:  # pragma: no cover - runtime only
            messagebox.showerror("Error", str(e))

    def _on_user(self) -> None:
        self._load_auth_headers()
        user_id = self.var_userid.get().strip()
        if not user_id:
            messagebox.showerror("Error", "User ID required")
            return
        try:
            self.user_data = fetch_worlds(
                user_id=user_id, limit=50, headers=self.headers
            )
            fetch_date = dt.datetime.now(dt.timezone.utc).strftime("%Y/%m/%d")
            for w in self.user_data:
                w["爬取日期"] = fetch_date
            with open(USER_FILE, "w", encoding="utf-8") as f:
                json.dump(self.user_data, f, ensure_ascii=False, indent=2)
            update_history(self.user_data)
            self.history = load_history()
            self._update_history_options()

            for item in self.user_tree.get_children():
                self.user_tree.delete(item)
            for w in self.user_data:
                row = record_row(w)
                self.user_tree.insert("", tk.END, values=row)
            self._create_world_tabs()
            self._update_dashboard()
            self.nb.select(self.tab_user)
        except RuntimeError as e:  # pragma: no cover - runtime only
            messagebox.showerror("HTTP Error", str(e))
        except Exception as e:  # pragma: no cover - runtime only
            messagebox.showerror("Error", str(e))

    def _update_tag_options(self) -> None:
        tags = set()
        for w in self.data:
            for t in w.get("tags", []):
                tags.add(t)
        self.box_tag["values"] = ["all"] + sorted(tags)
        self.var_tag.set("all")

    def _apply_filter(self) -> None:
        worlds = list(self.data)
        tag = self.var_tag.get()
        if tag != "all":
            worlds = [w for w in worlds if tag in w.get("tags", [])]
        if self.var_sort.get() == "latest":
            worlds.sort(key=lambda w: w.get("publicationDate", ""), reverse=True)
        else:
            worlds.sort(key=lambda w: w.get("visits", 0), reverse=True)
        self.filtered = worlds
        for item in self.tree.get_children():
            self.tree.delete(item)
        for w in self.filtered:
            name = w.get("name") or w.get("世界名稱")
            visits = w.get("visits") or w.get("瀏覽人次")
            world_id = w.get("id") or w.get("世界ID")
            self.tree.insert("", tk.END, values=(name, visits, world_id))
        self.nb.select(self.tab_list)

    def _update_history_options(self) -> None:
        self.hist_map: dict[str, str] = {}
        values = []
        for wid, recs in self.history.items():
            name = ""
            if isinstance(recs, list) and recs:
                name = recs[0].get("name", "")
            label = f"{name} ({wid})" if name else wid
            values.append(label)
            self.hist_map[label] = wid
        self.box_hist_world["values"] = values
        if values:
            self.var_hist_world.set(values[0])
            self._draw_history()

    def _draw_history(self, event=None) -> None:
        label = self.var_hist_world.get()
        world_id = getattr(self, "hist_map", {}).get(label, label)
        data = self.history.get(world_id, [])
        self.canvas.delete("all")
        if not data:
            return
        width = int(self.canvas.winfo_width() or 600)
        height = int(self.canvas.winfo_height() or 300)
        pad = 40
        times = [d["timestamp"] for d in data]
        min_t = min(times)
        max_t = max(times)
        if max_t == min_t:
            max_t += 1
        scale_x = width - 2 * pad
        scale_y = height - 2 * pad

        def xy(idx, val, max_val):
            x = pad + (times[idx] - min_t) / (max_t - min_t) * scale_x
            y = height - pad - min(val, max_val) / max_val * scale_y
            return x, y

        colors = {
            "visits": "blue",
            "favorites": "green",
            "heat": "red",
            "popularity": "purple",
        }
        limits = {
            "visits": 5000,
            "favorites": 5000,
            "heat": 10,
            "popularity": 10,
        }
        for key, color in colors.items():
            points = [xy(i, d.get(key, 0), limits[key]) for i, d in enumerate(data)]
            for a, b in zip(points, points[1:]):
                self.canvas.create_line(a[0], a[1], b[0], b[1], fill=color)
        # axes
        self.canvas.create_line(pad, height - pad, width - pad, height - pad)
        self.canvas.create_line(pad, pad, pad, height - pad)

    def _on_select_user_world(self, event=None) -> None:
        item = self.user_tree.focus()
        if not item:
            return
        values = self.user_tree.item(item, "values")
        if len(values) < 3:
            return
        world_id = values[2]
        self._draw_user_chart(world_id)

    def _draw_user_chart(self, world_id: str) -> None:
        data = self.history.get(world_id, [])
        self.user_canvas.delete("all")
        if not data:
            return
        width = int(self.user_canvas.winfo_width() or 600)
        height = int(self.user_canvas.winfo_height() or 200)
        pad = 40
        times = [d["timestamp"] for d in data]
        min_t = min(times)
        max_t = max(times)
        if max_t == min_t:
            max_t += 1
        scale_x = width - 2 * pad
        scale_y = height - 2 * pad

        def xy(idx, val, max_val):
            x = pad + (times[idx] - min_t) / (max_t - min_t) * scale_x
            y = height - pad - min(val, max_val) / max_val * scale_y
            return x, y

        colors = {"visits": "blue", "favorites": "green", "heat": "red", "popularity": "purple"}
        limits = {"visits": 5000, "favorites": 5000, "heat": 10, "popularity": 10}
        for key, color in colors.items():
            pts = [xy(i, d.get(key, 0), limits[key]) for i, d in enumerate(data)]
            for a, b in zip(pts, pts[1:]):
                self.user_canvas.create_line(a[0], a[1], b[0], b[1], fill=color)
        self.user_canvas.create_line(pad, height - pad, width - pad, height - pad)
        self.user_canvas.create_line(pad, pad, pad, height - pad)

    def _load_history_rows(self, world_id: str) -> list[dict]:
        """Return history rows for a world ID."""
        return list(self.history.get(world_id, []))

    def _draw_world_chart(self, canvas: tk.Canvas, world: dict) -> None:
        world_id = world.get("id") or world.get("worldId")
        data = self.history.get(world_id, [])
        canvas.delete("all")
        if not data:
            return

        width = int(canvas.winfo_width() or 600)
        height = int(canvas.winfo_height() or 200)
        pad = 40

        times = [d["timestamp"] for d in data]
        labs = _parse_date(world.get("labsPublicationDate"))
        pub = _parse_date(world.get("publicationDate"))
        update_times = []
        for d in data:
            u = _parse_date(d.get("updated_at"))
            if u:
                update_times.append(int(u.timestamp()))

        extra = [t for t in [labs, pub] if t]
        t_extra = [int(t.timestamp()) for t in extra] + update_times
        min_t = min([min(times)] + t_extra) if t_extra else min(times)
        max_t = max([max(times)] + t_extra) if t_extra else max(times)
        if max_t == min_t:
            max_t += 1

        scale_x = width - 2 * pad
        scale_y = height - 2 * pad

        def x_at(ts: int) -> float:
            return pad + (ts - min_t) / (max_t - min_t) * scale_x

        def y_val(val: float, limit: float) -> float:
            return height - pad - min(val, limit) / limit * scale_y

        colors = {
            "visits": "blue",
            "favorites": "green",
            "heat": "red",
            "popularity": "purple",
        }
        limits = {"visits": 10000, "favorites": 10000, "heat": 10, "popularity": 10}

        for key, color in colors.items():
            pts = []
            for rec in data:
                ts = rec["timestamp"]
                val = rec.get(key, 0) or 0
                pts.append((x_at(ts), y_val(val, limits[key])))
            for a, b in zip(pts, pts[1:]):
                canvas.create_line(a[0], a[1], b[0], b[1], fill=color)

        # event lines
        if labs:
            x = x_at(int(labs.timestamp()))
            canvas.create_line(x, pad, x, height - pad, fill="orange", dash=(4, 2))
        if pub:
            x = x_at(int(pub.timestamp()))
            canvas.create_line(x, pad, x, height - pad, fill="black", dash=(4, 2))
        for t in update_times:
            x = x_at(t)
            canvas.create_line(x, pad, x, height - pad, fill="gray", dash=(2, 2))

        canvas.create_line(pad, height - pad, width - pad, height - pad)
        canvas.create_line(pad, pad, pad, height - pad)
        canvas.create_line(width - pad, pad, width - pad, height - pad)

    def _create_world_tabs(self) -> None:
        """Create sub-tabs for each fetched user world with history."""
        if not hasattr(self, "detail_nb"):
            return
        # remove old tabs except the first (list tab)
        for tab_id in self.detail_nb.tabs()[1:]:
            self.detail_nb.forget(tab_id)


        unique: dict[str, dict] = {}
        for w in self.user_data:
            wid = w.get("世界ID") or w.get("worldId") or w.get("id")
            if wid:
                unique[wid] = w  # keep last occurrence

        for w in unique.values():
            frame = ttk.Frame(self.detail_nb)


            # dashboard table with a single metrics row
            dash = ttk.LabelFrame(frame, text="儀表板")
            dash.pack(fill=tk.X, padx=4, pady=2)
            dash_tree = ttk.Treeview(dash, columns=list(range(len(METRIC_COLS))), show="headings", height=2)
            for idx, col in enumerate(METRIC_COLS):
                dash_tree.heading(str(idx), text=col)
                dash_tree.column(str(idx), width=80, anchor="center")
            row = record_row(w)
            dash_tree.insert("", tk.END, values=row[1:])  # exclude fetch date
            dash_tree.pack(fill=tk.X, expand=True)

            # section 1: latest fetched info
            sec1 = ttk.LabelFrame(frame, text="本次資料")
            sec1.pack(fill=tk.X, padx=4, pady=2)
            info_tree = ttk.Treeview(sec1, columns=("k", "v"), show="headings", height=8)
            info_tree.heading("k", text="欄位")
            info_tree.heading("v", text="值")
            for key, val in w.items():
                info_tree.insert("", tk.END, values=(key, val))
            info_tree.pack(fill=tk.BOTH, expand=True)

            # section 2: history table from history JSON
            sec2 = ttk.LabelFrame(frame, text="歷史紀錄")
            sec2.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
            hist_tree = ttk.Treeview(sec2, show="headings")
            cols = [
                "timestamp",
                "visits",
                "favorites",
                "heat",
                "popularity",
                "updated_at",
                "created_at",
                "labs",
                "pub",
                "days_to_pub",
                "days_since_upd",
                "visits_per_day",
                "fav_per_day",
            ]
            hist_tree["columns"] = cols
            headers = [
                "時間",
                "人次",
                "收藏",
                "熱度",
                "熱門度",
                "更新",
                "上傳",
                "實驗室",
                "公開",
                "上傳到公開",
                "距離更新",
                "人次/天",
                "收藏/天",
            ]
            for c, h in zip(cols, headers):
                hist_tree.heading(c, text=h)
                hist_tree.column(c, width=80, anchor="center")
            rows = self._load_history_rows(w.get("id") or w.get("worldId"))
            for r in rows:
                ts = r["timestamp"]
                ts_dt = dt.datetime.fromtimestamp(ts, dt.timezone.utc)
                upd = _parse_date(r.get("updated_at"))
                created = _parse_date(r.get("created_at"))
                labs = _parse_date(r.get("labsPublicationDate"))
                pub = _parse_date(r.get("publicationDate"))
                ts_str = ts_dt.strftime("%Y/%m/%d")
                upd_str = upd.strftime("%Y/%m/%d") if upd else ""
                created_str = created.strftime("%Y/%m/%d") if created else ""
                labs_str = labs.strftime("%Y/%m/%d") if labs else ""
                pub_str = pub.strftime("%Y/%m/%d") if pub else ""
                days_to_pub = ""
                if pub and created:
                    days_to_pub = (pub - created).days
                elif pub and labs:
                    days_to_pub = (pub - labs).days
                days_since = (ts_dt - upd).days if upd else ""
                since_pub = (ts_dt - pub).days if pub else 0
                vpd = round((r.get("visits", 0) or 0) / since_pub, 2) if since_pub > 0 else ""
                fpd = round((r.get("favorites", 0) or 0) / since_pub, 2) if since_pub > 0 else ""
                hist_tree.insert(
                    "",
                    tk.END,
                    values=(
                        ts_str,
                        r.get("visits"),
                        r.get("favorites"),
                        r.get("heat"),
                        r.get("popularity"),
                        upd_str,
                        created_str,
                        labs_str,
                        pub_str,
                        days_to_pub,
                        days_since,
                        vpd,
                        fpd,
                    ),
                )
            hist_tree.pack(fill=tk.BOTH, expand=True)

            # section 3: chart
            sec3 = ttk.LabelFrame(frame, text="折線圖")
            sec3.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
            canvas = tk.Canvas(sec3, bg="white", height=200)
            canvas.pack(fill=tk.BOTH, expand=True)
            ttk.Label(sec3, text=LEGEND_TEXT).pack()
            self.after(100, lambda c=canvas, ww=w: self._draw_world_chart(c, ww))

            name = w.get("name") or w.get("世界名稱") or w.get("id")
            self.detail_nb.add(frame, text=str(name)[:15])

    def _update_dashboard(self) -> None:
        """Refresh the dashboard table and charts."""
        if not hasattr(self, "dash_tree"):
            return
        for item in self.dash_tree.get_children():
            self.dash_tree.delete(item)

        unique: dict[str, dict] = {}
        for w in self.user_data:
            wid = w.get("世界ID") or w.get("worldId") or w.get("id")
            if wid:
                unique[wid] = w
        for w in unique.values():
            row = record_row(w)
            self.dash_tree.insert("", tk.END, values=row[1:])

        for frame, _, _ in getattr(self, "chart_frames", []):
            frame.destroy()
        self.chart_frames = []
        for w in unique.values():
            frm = ttk.Frame(self.chart_container)
            canvas = tk.Canvas(frm, bg="white", width=240, height=180)
            canvas.pack(fill=tk.BOTH, expand=True)
            ttk.Label(frm, text=LEGEND_TEXT).pack()
            self.chart_frames.append((frm, canvas, w))
            self.after(100, lambda c=canvas, ww=w: self._draw_world_chart(c, ww))
        self._arrange_dashboard_charts()

    def _arrange_dashboard_charts(self, event=None) -> None:
        if not hasattr(self, "chart_frames"):
            return
        width = self.chart_container.winfo_width() if event is None else event.width
        cols = max(1, width // 260)
        for idx, (frm, _c, _w) in enumerate(self.chart_frames):
            frm.grid(row=idx // cols, column=idx % cols, padx=4, pady=4, sticky="nsew")
        for c in range(cols):
            self.chart_container.columnconfigure(c, weight=1)


def main() -> None:  # pragma: no cover - simple runtime entry
    try:
        app = WorldInfoUI()
    except tk.TclError as e:  # pragma: no cover - runtime only
        print("Failed to launch Tkinter UI:", e)
        return
    app.mainloop()


if __name__ == "__main__":
    main()
